from datetime import datetime as dt
from os.path import join
import openpyxl

TODAY = dt.now()
F_NAME_DATE_FORMAT = "%m-%d-%y"
TODAY_STR = dt.strftime(TODAY, F_NAME_DATE_FORMAT)
DST_F_PREFIX = f"Switch tracking numbers {TODAY_STR} QUICK"
DST_F_POSTFIX = ".xlsx"
DST_F_NAME = DST_F_PREFIX + DST_F_POSTFIX
DST_DIR = r'L:\Rollout\52648 FYE24 Network Refresh Switch\Coordinators\Switch shipment tracking'
DST_F_PATH = join(DST_DIR, DST_F_NAME)


def evaluate_formulas(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cell.value = cell.value


def copy_sheet_to_new_file(input_file, sheet_name, dst_file_path):
    try:
        # Load the Excel workbook and select the desired sheet
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook[sheet_name]

        # Evaluate the formulas and update cells with their values
        evaluate_formulas(sheet)

        # Create a new workbook and copy the contents
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = sheet_name

        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # Save the new workbook to a file
        new_workbook.save(dst_file_path)

        print(f'Contents from "{sheet_name}" in "{input_file}" copied to "{dst_file_path}" successfully.')
    except Exception as e:
        print(f'An error occurred: {str(e)}')


# Replace with your input and output file paths
input_file = r"C:\Users\beckett.mcfarland\Documents\Switch Report - Python.xlsx"  # Replace with your input file path
sheet_name = 'COVER'  # Replace with the sheet name you want to copy

copy_sheet_to_new_file(input_file, sheet_name, DST_F_PATH)
