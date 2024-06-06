import os
import time
from openpyxl import load_workbook

def copy_new_entries(source_file, source_sheet_name, destination_file, destination_sheet_name):
    # Check if the destination file exists, create if not
    if not os.path.exists(destination_file):
        wb = load_workbook(source_file)
        wb.save(destination_file)

    while True:
        # Load the source and destination workbooks
        source_wb = load_workbook(source_file)
        destination_wb = load_workbook(destination_file)
        
        source_sheet = source_wb[source_sheet_name]
        destination_sheet = destination_wb[destination_sheet_name]
        
        # Get the last row with data in the source sheet
        last_row_source = source_sheet.max_row

        # Check if new entries are available
        if last_row_source > destination_sheet.max_row:
            for row in range(destination_sheet.max_row, last_row_source):
                for col in range(1, source_sheet.max_column):
                    destination_sheet.cell(row=row, column=col, value=source_sheet.cell(row=row, column=col).value)
            destination_wb.save(destination_file)
            print("New entries copied.")
        
        source_wb.close()
        destination_wb.close()
        
        # Wait for a specified time interval before checking again
        time.sleep(60)  # Adjust this interval as needed

if __name__ == "__main__":
    source_file_path = r"C:\Users\beckett.mcfarland\Documents\Copy of Parcel Shipping.xlsm"  # Provide the path to the source Excel file
    source_sheet_name = "2023 Parcel Shipping"  # Specify the name of the source sheet
    
    destination_file_path = r"C:\Users\beckett.mcfarland\Documents\New Parcel.xlsx"  # Provide the path to the destination Excel file
    destination_sheet_name = "Sheet1"  # Specify the name of the destination sheet
    
    copy_new_entries(source_file_path, source_sheet_name, destination_file_path, destination_sheet_name)
