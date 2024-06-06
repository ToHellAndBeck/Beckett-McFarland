import win32com.client as win32
from datetime import datetime as dt
from os.path import abspath, basename, getmtime, join, exists
from os import PathLike, listdir
import pandas as pd
from openpyxl import Workbook, load_workbook

SCRIPT_PATH = abspath(__file__)
SCRIPT_DIR = SCRIPT_PATH.replace(basename(SCRIPT_PATH), "")
TODAY = dt.now()

# destination settings
F_NAME_DATE_FORMAT = "%m-%d-%y"
TODAY_STR = dt.strftime(TODAY, F_NAME_DATE_FORMAT)
DST_F_PREFIX = f"Switch Report {TODAY_STR} - Python"
DST_F_POSTFIX = ".xlsx"
DST_F_NAME = DST_F_PREFIX + DST_F_POSTFIX
DST_DIR = r"C:\Users\beckett.mcfarland\Documents"
DST_F_PATH = join(DST_DIR, DST_F_NAME)
DST_SHEET_NAME = "DAILY DUMP"
# source  settings
F_NAME_REQUIREMENTS = [
    "Master",
    "Switch",
    "Report",
]  # accept only files that contain these
F_TYPES_ALLOWED = [".xlsx", ".xlsm"]  # accept only files that end in these
SRC_F_DIR = r"C:\Users\beckett.mcfarland\Documents\output_excel_files"# change this to the directory the source file is located
SRC_SHEET_NAME = "Conference Call Switch"
MIN_SRC_COLUMNS = 5


def file_meets_requirements(f_name: str):
    for ending in F_TYPES_ALLOWED:
        if f_name.endswith(ending):
            for requirement in F_NAME_REQUIREMENTS:
                if requirement not in f_name:
                    return False
            return True


def get_most_recent_file(f_dir: PathLike):
    most_recent_name = None
    most_recent_time = None
    for f_name in listdir(f_dir):
        if file_meets_requirements(f_name):
            f_path = join(f_dir, f_name)
            f_time = getmtime(f_path)
            if not most_recent_name:
                most_recent_name = f_name
                most_recent_time = f_time
            else:
                if f_time > most_recent_time:
                    most_recent_name = f_name
                    most_recent_time = f_time
    return join(f_dir, most_recent_name)


file_to_wrk = get_most_recent_file(SRC_F_DIR)

# Use win32com to open and save the file (for any pre-processing in Excel)
excel = win32.gencache.EnsureDispatch('Excel.Application')
workbook = excel.Workbooks.Open(file_to_wrk)
workbook.Save()
workbook.Close()
excel.Quit()

# Now load the workbook with openpyxl
src_wb = load_workbook(file_to_wrk, data_only=False)
srcwb2: dict[str, pd.DataFrame] = pd.read_excel(
    file_to_wrk, sheet_name=[SRC_SHEET_NAME], index_col=None
)
src_sheet_2 = srcwb2[SRC_SHEET_NAME]

# Check if destination file exists and create it if it doesn't
if not exists(DST_F_PATH):
    dst_wb = Workbook()
    dst_ws = dst_wb.worksheets[0]
    dst_ws.title = DST_SHEET_NAME
    dst_wb.save(DST_F_PATH)
    dst_wb.close()

# Write to the destination file
with pd.ExcelWriter(DST_F_PATH, mode="a", if_sheet_exists="overlay") as writer:
    src_sheet_2.to_excel(writer, sheet_name=DST_SHEET_NAME)
