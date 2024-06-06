from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
import time
import os
import openpyxl
from datetime import datetime, timedelta
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains

# Set download directory
download_dir = r"C:\Users\beckett.mcfarland\Downloads\Survey Pull"  # Replace with the path to your download directory
chrome_options = Options()
chrome_options.add_experimental_option("prefs", {
  "download.default_directory": download_dir,
  "download.prompt_for_download": False,  # To automatically save files to the specified directory
  "download.directory_upgrade": True,
  "safebrowsing.enabled": True
})
'''
def select_option_from_custom_dropdown_survey_name(driver, dropdown_xpath, option_text):
    try:
        # Click the dropdown to open it
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, dropdown_xpath)))
        dropdown.click()

        # Wait for the dropdown options to be clickable
        option_xpath = '//*[contains(text(), "52648 Switch Refresh WO")]'
        option = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, option_xpath)))

        # Scroll into view and click the option
        driver.execute_script("arguments[0].scrollIntoView(true);", option)
        option.click()
    except Exception as e:
        print(f"Error in selecting option from dropdown: {e}")
'''
def select_option_from_custom_dropdown_name(driver, dropdown_xpath):
    try:
        # Click the dropdown to open it
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, dropdown_xpath)))
        dropdown.click()

        # Wait a moment for the dropdown to open fully
        time.sleep(1)  # Adjust the sleep time as needed

        # Use ActionChains to send the '5' key three times to the dropdown
        actions = ActionChains(driver)
        '''
        for _ in range(3):
            actions.send_keys('5').perform()
            time.sleep(1)  # Brief pause between key presses; adjust as needed
        '''
        actions.send_keys('5').perform()
        actions.send_keys('5').perform()
        actions.send_keys(Keys.RETURN).perform()
        # Add any additional actions here if needed, like selecting the option

    except Exception as e:
        print(f"Error in selecting option from dropdown: {e}")

def select_option_from_custom_dropdown_version(driver, dropdown_xpath):
    try:
        # Click the dropdown to open it
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, dropdown_xpath)))
        dropdown.click()

        # Wait a moment for the dropdown to open fully
        time.sleep(1)  # Adjust the sleep time as needed

        # Use ActionChains to send the '5' key three times to the dropdown
        actions = ActionChains(driver)
        '''
        for _ in range(3):
            actions.send_keys('5').perform()
            time.sleep(1)  # Brief pause between key presses; adjust as needed
        '''
        actions.send_keys('5').perform()
        actions.send_keys('5').perform()
        actions.send_keys('5').perform()
        actions.send_keys(Keys.RETURN).perform()
        # Add any additional actions here if needed, like selecting the option

    except Exception as e:
        print(f"Error in selecting option from dropdown: {e}")

def read_numbers_from_excel(file_path, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    numbers = []
    for cell in sheet[column]:
        if cell.value is not None:
            numbers.append(str(cell.value))
    return numbers

def write_to_excel(file_path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, (text1, text2) in enumerate(data, start=1):
        sheet.cell(row=i, column=1).value = text1
        sheet.cell(row=i, column=2).value = text2
    workbook.save(file_path)
def input_text_in_search_box(driver, search_box_xpath, number):
    try:
        search_box = driver.find_element(By.XPATH, search_box_xpath)
        search_box.clear()
        search_box.send_keys(number)
        search_box.send_keys(Keys.ENTER)
        
        # Wait for the element containing the exact number to be visible
        number_xpath = f'//*[contains(text(), "{number.strip()}")]'
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, number_xpath))
        )
    except Exception as e:
        print(f"Error inputting text in search box: {e}")

def find_row_with_number(driver, number, table_xpath):
    try:
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        for row in rows:
            cells = row.find_elements(By.XPATH, "./td")
            for i, cell in enumerate(cells):
                if cell.text.strip() == number.strip():
                    return row, i  # Return the row element and the index of the found cell
        return None, None
    except Exception as e:
        print(f"Error finding row with number {number}: {e}")
        return None, None

def get_text_from_next_column(row_element, cell_index):
    try:
        next_column_xpath = f'./td[{cell_index + 2}]'  # Increment by 2 to get the next cell
        return row_element.find_element(By.XPATH, next_column_xpath).text
    except Exception as e:
        print(f"Error getting text from next column: {e}")
        return None

def wait_for_element_to_be_clickable(driver, xpath):
    try:
        button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        return button
    except Exception as e:
        print(xpath, e)
 
def click_button(driver, xpath):
    try:
        button = wait_for_element_to_be_clickable(driver, xpath)
        if button:
            button.click()
    except Exception as e:
        print(xpath, e)
 
def input_text(driver, element_id, text):
    try:
        # Find the email input field by its ID
        email_input = driver.find_element(By.ID, element_id)
        # Clear the input field if needed
        email_input.clear()
        # Input the email address
        email_input.send_keys(text)
        email_input.send_keys(Keys.ENTER)
    except Exception as e:
        print(f"An error occurred: {e}")

def get_text_from_xpath(driver, xpath):
    try:
        element = driver.find_element(By.XPATH, xpath)
        return element.text
    except Exception as e:
        print(f"Error finding element with xpath {xpath}: {e}")
        return None 
def input_number_and_search(driver, search_box_xpath, number):
    try:
        search_box = driver.find_element(By.XPATH, search_box_xpath)
        search_box.clear()
        search_box.send_keys(number)
        search_box.send_keys(Keys.ENTER)
        
        # More robust wait: wait until the table updates with the search results
        WebDriverWait(driver, 10).until(
            EC.text_to_be_present_in_element((By.XPATH, table_xpath + "/tbody/tr[1]/td[3]"), number.strip())
        )
    except Exception as e:
        print(f"Error inputting number in search box: {e}")

def find_matching_row_and_get_data(driver, number, table_xpath):
    try:
        number_str = str(number).strip()  # Ensure the number is treated as a string
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        for row in rows:
            cell_text = row.find_element(By.XPATH, "./td[3]").text.strip()
            if cell_text == number_str:  # Compare as strings
                cell_text_1 = cell_text
                cell_text_2 = row.find_element(By.XPATH, "./td[4]/a").text.strip()
                return cell_text_1, cell_text_2
        return None, None
    except Exception as e:
        print(f"Error finding matching row: {e}")
        return None, None


def wait_for_file_download(download_dir, file_prefix, timeout=3000):
    """
    Waits for a file to be downloaded in the specified directory.
    :param download_dir: Directory where Chrome downloads files.
    :param file_prefix: Prefix or part of the filename to identify the correct file.
    :param timeout: Maximum time to wait for the file to download, in seconds.
    """
    end_time = time.time() + timeout
    while True:
        for filename in os.listdir(download_dir):
            if filename.startswith(file_prefix):
                return os.path.join(download_dir, filename)
        if time.time() > end_time:
            raise TimeoutError("File download timed out.")
        time.sleep(5)  # Check every 5 seconds

# Initialize the Chrome driver
# chrome_options=Options()
# chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=chrome_options)
url = 'https://trust.wachter.com/Rollout/MultiSiteManagement.aspx?JobId=942451'
employee_dropdown_xpath = '//*[@id="ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input"]'
rcb_list_xpath = '/html/body/form/div[1]/div/div/ul'
email_element_id = 'i0116'
dropdown_input_id = 'ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input'
driver.get(url)
login_button_xpath = '//*[@id="slide-5-layer-8"]'
click_button(driver, login_button_xpath)
wachter_login_button_xpath = '//*[@id="ctl00_cphMain_ssoLoginControl_btnSingleSignOnLogin"]'
click_button(driver, wachter_login_button_xpath)
next_button_path = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, next_button_path)
email_address = 'beckett.mcfarland@wachter.com'
input_text(driver, email_element_id, email_address)
sign_in_button_xpath = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, sign_in_button_xpath)
password_element_id = 'i0118'
password = 'Ozymandias99!'
input_text(driver, password_element_id, password)
stay_signed_in_button_xpath = '//*[@id="idSIButton9"]'
time.sleep(3)
click_button(driver, stay_signed_in_button_xpath)
time.sleep(8)
survey_button_xpath='//*[@id="ctl00_cphMain_tabSurveys"]/a'
click_button(driver,survey_button_xpath)
time.sleep(3)
download_csv_button='//*[@class="fas fa-file-csv wa-white"]'
click_button(driver,download_csv_button)
time.sleep(3)
dropdown1_xpath = '/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/div/div[1]/kendo-dropdownlist/span/span[2]/span'
dropdown2_xpath='//div[2]/kendo-dropdownlist/span/span'
time.sleep(3)
select_option_from_custom_dropdown_name(driver, dropdown1_xpath)
time.sleep(2)
select_option_from_custom_dropdown_version(driver, dropdown2_xpath)
export_xpath='/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/div/div[3]/button[2]'
click_button(driver,export_xpath)
time.sleep(30)
file_prefix = 'SurveyResponses'  # The part of the filename that you know in advance
downloaded_file_path = wait_for_file_download(download_dir, file_prefix)

""" 
dropdown1_option_text = "52648 Switch Refresh WO - v12 - 8/1/2023"
= '/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/div/div[2]/kendo-dropdownlist/span/span[1]'
dropdown2_option_text = "52648 Switch Refresh WO - V12"
select_option_from_custom_dropdown_version(driver, dropdown2_xpath, dropdown2_option_text)
select_option_from_custom_dropdown_survey_name(driver, dropdown1_xpath, dropdown1_option_text)

"""

