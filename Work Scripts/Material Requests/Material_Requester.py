import openpyxl

table_data = [
["Time", "Sender", "Team", "Site", "Item", "Quantity"],
["10:42 AM", "CJ Ewell", "Switch team 8", "882", "16 sfp 10 gig, 2 stacking cables", "N/A"],
["10:42 AM", "CJ Ewell", "Switch team 17", "872", "8 stacking cables, 2 bags of cage nuts", "N/A"],
["10:43 AM", "CJ Ewell", "switch team 40", "3729", "4 switches, 4 sets of stacking cables, 6 sfp 10 gig", "N/A"],
["10:43 AM", "CJ Ewell", "switch team 24", "1458", "2 stacking cables", "N/A"],
["10:43 AM", "CJ Ewell", "switch team 55", "2726", "1 NHOST 1 and 2 shipped out", "N/A"]






]

import datetime
output_directory = r"C:\Users\beckett.mcfarland\Desktop\Beckett-McFarland\Work Scripts\Material Requests\Output\\"
today = datetime.datetime.now().date()
key_coordinates = {"Site": (3, 9), "Item": (17, 5)}
headers = []
template_location = r"C:\Users\beckett.mcfarland\Desktop\Beckett-McFarland\Work Scripts\Material Requests\Template\Walmart Shipping Request Template (Personal).xlsx"
wb = openpyxl.load_workbook(template_location)
ws = wb["Sheet1"]

for row_idx, row in enumerate(table_data):

    if row_idx == 0:
        for column in row:
            headers.append(column)
    else:
        name_list = []
        for column_idx, column_value in enumerate(row):
            if column_idx == 2 or column_idx == 3:
                name_list.append(column_value)

            key = headers[column_idx]
            position = key_coordinates.get(key)
            if position:
                row = position[0]
                col = position[1]
                if key == "Item":
                    # Split items separated by commas and handle numbers at the beginning
                    items = column_value.split(", ")
                    for i, item in enumerate(items):
                        item_parts = item.split(" ", 1)
                        if len(item_parts) == 2 and item_parts[0].isdigit():
                            ws.cell(row + i, col - 1).value = int(item_parts[0])
                            ws.cell(row + i, col).value = item_parts[1]
                        else:
                            # Check if the item part is a number and convert to int
                            item_value = item_parts[0] if not item_parts[0].isdigit() else int(item_parts[0])
                            ws.cell(row + i, col).value = item_value
                else:
                    ws.cell(row, col).value = column_value

        name_list.append(str(today))
        file_name = output_directory + "-".join(name_list) + ".xlsx"
        wb.save(file_name)
