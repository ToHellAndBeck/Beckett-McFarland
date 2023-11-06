import os
import openpyxl

# Set the folder path where your Excel files are located
folder_path = r'L:\Rollout\52642 FYE24 Network Refresh Fiber\Shipping'

# Initialize an empty list to store extracted data
data_list = []

# Iterate through files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):  # Assuming all files are in the .xlsx format
        file_path = os.path.join(folder_path, filename)

        # Extract the date and team number from the file name
        file_parts = filename.split(" Team ")
        if len(file_parts) == 2:
            date = file_parts[0].split()[-1]
            team_number = file_parts[1][:-1]  # Remove the closing bracket from the team number
        else:
            # If there is no team number after "Team," assume it's part of the date
            date_and_more = filename.split(" Team ")[0].split()[-1]
            date = date_and_more if len(date_and_more) > 0 else "N/A"
            team_number = "N/A"

        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path)

        # Access the specific sheet in the Excel file (change 'Sheet1' to your sheet name)
        sheet = workbook['Sheet1']

        # Extract data from specific cells
        store_number = sheet['I3'].value
        job_number = sheet['C15'].value

        # Extract items from columns D (4) and E (5) for the range E17 to E32
        items = [str(sheet.cell(row=row, column=4).value) + " " + str(sheet.cell(row=row, column=5).value) for row in range(17, 33) if sheet.cell(row=row, column=4).value is not None]

        # Store the extracted data in a dictionary
        data_dict = {
            'Store Number': store_number,
            'Job Number': job_number,
            'Team Number': team_number,
            'Date': date,
            'Items': items if items else None
        }

        data_list.append(data_dict)

# Create a text document to compile the data
output_file = r'C:\Users\beckett.mcfarland\Documents\txt_files\output(fiber).txt'

# Write the extracted data to the text document
with open(output_file, 'w') as text_file:
    for data in data_list:
        text_file.write(f"Store #: {data['Store Number']}\n")
        text_file.write(f"Job #: {data['Job Number']}\n")
        text_file.write(f"Team #: {data['Team Number']}\n")
        text_file.write(f"Date: {data['Date']}\n")
        if data['Items'] is not None:
            text_file.write("Items:\n")
            for item in data['Items']:
                text_file.write(f"- {item}\n")
        text_file.write("\n")

print(f"Data extracted and compiled into {output_file}")
