import os
import time
from openpyxl import load_workbook

def copy_new_entries(source_file, destination_file):
    # Check if the destination file exists, create if not
    if not os.path.exists(destination_file):
        wb = load_workbook(source_file)
        wb.save(destination_file)

    while True:
        # Load the source and destination workbooks
        source_wb = load_workbook(source_file)
        destination_wb = load_workbook(destination_file)
        
        source_sheet = source_wb.active
        destination_sheet = destination_wb.active
        
        # Get the last row with data in the source sheet
        last_row_source = source_sheet.max_row

        # Check if new entries are available
        if last_row_source > destination_sheet.max_row:
            for row in range(destination_sheet.max_row + 1, last_row_source + 1):
                for col in range(1, source_sheet.max_column + 1):
                    destination_sheet.cell(row=row, column=col, value=source_sheet.cell(row=row, column=col).value)
            destination_wb.save(destination_file)
            print("New entries copied.")
        
        source_wb.close()
        destination_wb.close()
        
        # Wait for a specified time interval before checking again
        time.sleep(60)  # Adjust this interval as needed

if __name__ == "__main__":
    source_file_path = "L:\Warehouse\Freight shipping.xlsx"  # Provide the path to the source Excel file
    destination_file_path = "C:\Users\beckett.mcfarland\Documents\Switch tracking numbers 8-25-23.xlsx"  # Provide the path to the destination Excel file
    
    copy_new_entries(source_file_path, destination_file_path)
