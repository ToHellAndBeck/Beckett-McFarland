import openpyxl

# Load the source workbook
source_workbook = openpyxl.load_workbook('source.xlsx')
source_sheet = source_workbook.active

# Create a new workbook for the destination
destination_workbook = openpyxl.Workbook()
destination_sheet = destination_workbook.active

# Copy data from source to destination
for row in source_sheet.iter_rows():
    for cell in row:
        destination_sheet[cell.coordinate].value = cell.value

# Save the destination workbook
destination_workbook.save('destination.xlsx')

# Close the workbooks
source_workbook.close()
destination_workbook.close()

print("Data copied and saved to destination.xlsx")
