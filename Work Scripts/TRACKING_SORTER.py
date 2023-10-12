import openpyxl

def read_column_values(input_file, sheet_name, column):
    # Load the Excel workbook and select the sheet
    workbook = openpyxl.load_workbook(input_file, data_only=True)  # Use data_only to read values instead of formulas
    sheet = workbook[sheet_name]

    # Read values from the specified column (excluding formulas and #N/A)
    values = [sheet.cell(row=row, column=column).value for row in range(1, sheet.max_row + 1) if sheet.cell(row=row, column=column).value is not None and sheet.cell(row=row, column=column).value != "#N/A"]

    # Close the workbook
    workbook.close()

    return values

def write_values_to_new_file(values, output_file):
    # Create a new workbook and select the sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Split values into groups of 30 and write them to different columns
    for i, chunk in enumerate(zip(*[iter(values)]*30), start=1):
        for j, value in enumerate(chunk, start=1):
            sheet.cell(row=j, column=i).value = value

    # Save the new workbook to the specified output file
    workbook.save(output_file)
    workbook.close()

if __name__ == "__main__":
    # Input file, sheet name, column to read (A=1, B=2, etc.)
    input_file = r"C:\Users\beckett.mcfarland\Documents\Switch Report - Python.xlsx"
    sheet_name = "COVER"
    column_to_read = 15  # Assuming we're reading the first column (A)

    # Output file for grouped values
    output_file = r"C:\Users\beckett.mcfarland\Documents\output_excel_files\TrackingSort.xlsx"

    # Read values from the specified column excluding #N/A
    values = read_column_values(input_file, sheet_name, column_to_read)

    # Write grouped values to a new Excel file
    write_values_to_new_file(values, output_file)

    print("Values grouped and written to", output_file)
