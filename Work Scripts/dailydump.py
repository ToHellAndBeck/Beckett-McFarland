<<<<<<< HEAD
from os import listdir, makedirs
from os.path import abspath, basename, join, exists, getmtime
from openpyxl import load_workbook
from datetime import datetime as dt
import pandas as pd
=======
import xlrd
import openpyxl
import os
filepath = r"C:\Users\beckett.mcfarland\Documents"
# Load the source workbook using xlrd
source_workbook = xlrd.open_workbook('copy of Master Switch Report 8-27-23.xlsb')

# Specify the source and destination sheets by name
source_sheet_name = 'Conference Call Switch'
destination_workbook = openpyxl.load_workbook('Switch tracking numbers 8-25-23.xlsb')
destination_sheet_name = 'DAILY DUMP'

# Get the source sheet using xlrd
source_sheet = source_workbook.sheet_by_name(source_sheet_name)
>>>>>>> 4082a405aa92e157feddc7615c2b4d4431bf1b55

# Get the destination sheet using openpyxl
destination_sheet = destination_workbook[destination_sheet_name]
print(source_sheet.values)
# Copy data from source to destination
# for row_index in range(source_sheet.nrows):
#    for col_index in range(source_sheet.ncols):
#        cell_value = source_sheet.cell(row_index, col_index).value
#       destination_sheet.cell(row=row_index + 1, column=col_index + 1).value = cell_value

<<<<<<< HEAD
# destination settings
F_NAME_DATE_FORMAT = "%m-%d-%y"
TODAY_STR = dt.strftime(TODAY, F_NAME_DATE_FORMAT)
DST_F_PREFIX = f"Switch Report {TODAY_STR} - Python"
DST_F_POSTFIX = ".xlsx"
DST_F_NAME = DST_F_PREFIX + DST_F_POSTFIX
DST_DIR = SCRIPT_DIR
DST_F_PATH = join(DST_DIR, r"C:\Users\beckett.mcfarland\Documents\Switch Report - Python.xlsx")
DST_SHEET_NAME = "DAILY DUMP"
# source  settings
F_NAME_REQUIREMENTS = [
    "Master",
    "Switch",
    "Report",
]  # accept only files that contain these
F_TYPES_ALLOWED = [".xlsx", ".xlsm"]  # accept only files that end in these
SRC_F_DIR = r'C:\Users\beckett.mcfarland\Documents\output_excel_files'  # change this to the directory the source file is located
SRC_SHEET_NAME = "Conference Call Switch"
MIN_SRC_COLUMNS = 5
=======
# Save the destination workbook
destination_workbook.save('destination.xlsb')
>>>>>>> 4082a405aa92e157feddc7615c2b4d4431bf1b55

# Close the destination workbook (openpyxl handles source workbook closing)
destination_workbook.close()

<<<<<<< HEAD
def get_most_recent_file(f_dir: str):
    most_recent_name = None
    most_recent_time = None
    for f_name in listdir(f_dir):
        if file_meets_requirements(f_name):
            f_path = join(f_dir, f_name)
            f_time = getmtime(f_path)
            if not most_recent_name:
                most_recent_name = f_name
                most_recent_time = f_time
            else:
                if f_time > most_recent_time:
                    most_recent_name = f_name
                    most_recent_time = f_time

    if most_recent_name is not None:
        return join(f_dir, most_recent_name)
    else:
        print("No file found that meets the requirements.")
        return None  # Return a meaningful value when no file is found

file_to_wrk = get_most_recent_file(SRC_F_DIR)
src_wb = load_workbook(file_to_wrk, data_only=True)

if not exists(DST_F_PATH):
    print(f"The file '{DST_F_PATH}' does not exist.")
    exit()

# Load the existing Excel workbook
dst_wb = load_workbook(DST_F_PATH)

# Check if the destination sheet already exists
if DST_SHEET_NAME in dst_wb.sheetnames:
    # Remove the existing sheet
    dst_wb.remove(dst_wb[DST_SHEET_NAME])

# Get the source sheet
src_sheet = src_wb[SRC_SHEET_NAME]

# Create a new sheet in the destination workbook
dst_sheet = dst_wb.create_sheet(title=DST_SHEET_NAME)

# Copy data from source sheet to destination sheet
for row in src_sheet.iter_rows():
    for cell in row:
        dst_sheet[cell.coordinate].value = cell.value

# Save the destination workbook
dst_wb.save(DST_F_PATH)
print(f"The sheet '{DST_SHEET_NAME}' has been overwritten in '{DST_F_PATH}'.")
=======
print("Data copied and saved to destination.xlsb")
>>>>>>> 4082a405aa92e157feddc7615c2b4d4431bf1b55
