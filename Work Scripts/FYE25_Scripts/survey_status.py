from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
from datetime import datetime, timedelta

def get_modified_number(number):
    """
    Prefixes zeros to the number based on its length.
    2 digits -> 2 zeros, 3 digits -> 1 zero.
    """
    if len(number) == 2:
        return '00' + number
    elif len(number) == 3:
        return '0' + number
    else:
        return number
    
def read_numbers_from_excel(file_path, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    numbers = []
    for cell in sheet[column]:
        if cell.value is not None:
            numbers.append(str(cell.value))
    return numbers

def write_to_excel(file_path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, (text1, text2) in enumerate(data, start=1):
        sheet.cell(row=i, column=1).value = text1
        sheet.cell(row=i, column=2).value = text2
    workbook.save(file_path)
def input_text_in_search_box(driver, search_box_xpath, number):
    try:
        search_box = driver.find_element(By.XPATH, search_box_xpath)
        search_box.clear()
        search_box.send_keys(number)
        search_box.send_keys(Keys.ENTER)
        
        # Wait for the element containing the exact number to be visible
        number_xpath = f'//*[contains(text(), "{number.strip()}")]'
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, number_xpath))
        )
    except Exception as e:
        print(f"Error inputting text in search box: {e}")

def find_row_with_number(driver, number, table_xpath):
    try:
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        for row in rows:
            cells = row.find_elements(By.XPATH, "./td")
            for i, cell in enumerate(cells):
                if cell.text.strip() == number.strip():
                    return row, i  # Return the row element and the index of the found cell
        return None, None
    except Exception as e:
        print(f"Error finding row with number {number}: {e}")
        return None, None

def get_text_from_next_column(row_element, cell_index):
    try:
        next_column_xpath = f'./td[{cell_index + 2}]'  # Increment by 2 to get the next cell
        return row_element.find_element(By.XPATH, next_column_xpath).text
    except Exception as e:
        print(f"Error getting text from next column: {e}")
        return None

def wait_for_element_to_be_clickable(driver, xpath):
    try:
        button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        return button
    except Exception as e:
        print(xpath, e)
 
def click_button(driver, xpath):
    try:
        button = wait_for_element_to_be_clickable(driver, xpath)
        if button:
            button.click()
    except Exception as e:
        print(xpath, e)
 
def input_text(driver, element_id, text):
    try:
        # Find the email input field by its ID
        email_input = driver.find_element(By.ID, element_id)
        # Clear the input field if needed
        email_input.clear()
        # Input the email address
        email_input.send_keys(text)
        email_input.send_keys(Keys.ENTER)
    except Exception as e:
        print(f"An error occurred: {e}")

def get_text_from_xpath(driver, xpath):
    try:
        element = driver.find_element(By.XPATH, xpath)
        return element.text
    except Exception as e:
        print(f"Error finding element with xpath {xpath}: {e}")
        return None 
    
def find_all_matching_rows_and_get_data(driver, number, table_xpath, specific_text):
    try:
        number_str = str(number).strip()  # Ensure the number is treated as a string
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        matched_data = []
        for row_index, row in enumerate(rows, start=1):
            site_number_text = row.find_element(By.XPATH, f"{table_xpath}/tbody/tr[{row_index}]/td[3]").text.strip()
            specific_column_text = row.find_element(By.XPATH, f"{table_xpath}/tbody/tr[{row_index}]/td[5]").text.strip()
            if site_number_text == number_str and specific_column_text == specific_text:  # Compare as strings
                cell_text_1 = site_number_text
                cell_text_2 = row.find_element(By.XPATH, f"{table_xpath}/tbody/tr[{row_index}]/td[4]/a").text.strip()
                matched_data.append((cell_text_1, cell_text_2))
        return matched_data
    except Exception as e:
        print(f"Error finding matching rows: {e}")
        return []

    
def input_number_and_search(driver, search_box_xpath, number):
    try:
        search_box = driver.find_element(By.XPATH, search_box_xpath)
        search_box.clear()
        search_box.send_keys(number)
        search_box.send_keys(Keys.ENTER)
        
        # More robust wait: wait until the table updates with the search results
        WebDriverWait(driver, 10).until(
            EC.text_to_be_present_in_element((By.XPATH, table_xpath + "/tbody/tr[1]/td[3]"), number.strip())
        )
    except Exception as e:
        print(f"Error inputting number in search box: {e}")

def find_matching_row_and_get_data(driver, number, table_xpath):
    try:
        number_str = str(number).strip()  # Ensure the number is treated as a string
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        for row in rows:
            cell_text = row.find_element(By.XPATH, "./td[3]").text.strip()
            if cell_text == number_str:  # Compare as strings
                cell_text_1 = cell_text
                cell_text_2 = row.find_element(By.XPATH, "./td[4]/a").text.strip()
                return cell_text_1, cell_text_2
        return None, None
    except Exception as e:
        print(f"Error finding matching row: {e}")
        return None, None

# Initialize the Chrome driver
# chrome_options=Options()
# chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))#,options=chrome_options
url = 'https://trust.wachter.com/Rollout/MultiSiteManagement.aspx?JobId=1012333'
employee_dropdown_xpath = '//*[@id="ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input"]'
rcb_list_xpath = '/html/body/form/div[1]/div/div/ul'
email_element_id = 'i0116'
dropdown_input_id = 'ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input'
driver.get(url)
login_button_xpath = '//*[@id="slide-5-layer-8"]'
click_button(driver, login_button_xpath)
wachter_login_button_xpath = '//*[@id="ctl00_cphMain_ssoLoginControl_btnSingleSignOnLogin"]'
click_button(driver, wachter_login_button_xpath)
next_button_path = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, next_button_path)
email_address = 'beckett.mcfarland@wachter.com'
input_text(driver, email_element_id, email_address)
sign_in_button_xpath = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, sign_in_button_xpath)
password_element_id = 'i0118'
password = 'Ozymandias99!'
input_text(driver, password_element_id, password)
stay_signed_in_button_xpath = '//*[@id="idSIButton9"]'
click_button(driver, stay_signed_in_button_xpath)
time.sleep(15)
survey_button_xpath='//*[@id="ctl00_cphMain_tabSurveys"]/a'
click_button(driver,survey_button_xpath)
search_box="/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/span/kendo-grid/div/div/table/thead/tr[2]/th[3]/text-box-filter/span/input"
text_xpath = '/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/span/kendo-grid/kendo-grid-list/div/div[1]/table/tbody/tr[1]/td[3]'
time.sleep(5)
# Copying text from the specified XPath
# copied_text = get_text_from_xpath(driver, text_xpath)
# print("Copied text:", copied_text)

time.sleep(5)
numbers = read_numbers_from_excel(r"L:\Rollout\97820 WM - Switch-Fiber FYE25 Network Refresh - CJE\Coordinator\97820_Scripts\Surveys\survey_checker\survey_input\survey_input.xlsx", "A")  # Replace with your Excel file path and column

text_xpath_1 = '/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/span/kendo-grid/kendo-grid-list/div/div[1]/table/tbody/tr[1]/td[3]'
text_xpath_2 = '/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/span/kendo-grid/kendo-grid-list/div/div[1]/table/tbody/tr[1]/td[4]/a'

# Iterate through the numbers and perform search and copy actions
copied_data = []
search_box_xpath = "/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/span/kendo-grid/div/div/table/thead/tr[2]/th[3]/text-box-filter/span/input"
table_xpath = '/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/span/kendo-grid/kendo-grid-list/div/div[1]/table'
specific_search_box_xpath = "/html/body/form/div[8]/div/div/div[1]/rollout-management-app/div/div/survey-grid/spinner-container/div/span/kendo-grid/div/div/table/thead/tr[2]/th[5]/text-box-filter/span/input"

specific_text_to_enter="97820 Walmart Fiber and Switch"

input_text_in_search_box(driver, specific_search_box_xpath, specific_text_to_enter)


specific_text_to_check = "97820 Walmart Fiber and Switch"

for number in numbers:
    time.sleep(2)
    print(f"Searching for number: {number}")
    input_number_and_search(driver, search_box_xpath, number)
    matched_data = find_all_matching_rows_and_get_data(driver, number, table_xpath, specific_text_to_check)

    if not matched_data:
        # Retry with modified number
        modified_number = get_modified_number(number)
        print(f"Retrying with modified number: {modified_number}")
        input_number_and_search(driver, search_box_xpath, modified_number)
        matched_data = find_all_matching_rows_and_get_data(driver, modified_number, table_xpath, specific_text_to_check)

    if matched_data:
        for cell_text_1, cell_text_2 in matched_data:
            print(f"Copied texts for {number.strip()}: {cell_text_1}, {cell_text_2}")
            copied_data.append((cell_text_1, cell_text_2))
    else:
        print(f"Row with number {number.strip()} not found.")
        copied_data.append((number, "couldn't locate"))

# Write the results to a new Excel file
write_to_excel(r"L:\Rollout\97820 WM - Switch-Fiber FYE25 Network Refresh - CJE\Coordinator\97820_Scripts\Surveys\survey_checker\survey_outputs\survey_output.xlsx", copied_data)
