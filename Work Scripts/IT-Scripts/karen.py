
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta

def get_modified_number(number):
    """
    Prefixes zeros to the number based on its length.
    2 digits -> 2 zeros, 3 digits -> 1 zero.
    """
    if len(number) == 2:
        return '00' + number
    elif len(number) == 3:
        return '0' + number
    else:
        return number
    
def read_numbers_from_excel(file_path, column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    numbers = []
    for cell in sheet[column]:
        if cell.value is not None:
            numbers.append(str(cell.value))
    return numbers

def write_to_excel(file_path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, (text1, text2) in enumerate(data, start=1):
        sheet.cell(row=i, column=1).value = text1
        sheet.cell(row=i, column=2).value = text2
    workbook.save(file_path)
def input_text_in_search_box(driver, search_box_xpath, number):
    try:
        search_box = driver.find_element(By.XPATH, search_box_xpath)
        search_box.clear()
        search_box.send_keys(number)
        search_box.send_keys(Keys.ENTER)
        
        # Wait for the element containing the exact number to be visible
        number_xpath = f'//*[contains(text(), "{number.strip()}")]'
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, number_xpath))
        )
    except Exception as e:
        print(f"Error inputting text in search box: {e}")

def find_row_with_number(driver, number, table_xpath):
    try:
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        for row in rows:
            cells = row.find_elements(By.XPATH, "./td")
            for i, cell in enumerate(cells):
                if cell.text.strip() == number.strip():
                    return row, i  # Return the row element and the index of the found cell
        return None, None
    except Exception as e:
        print(f"Error finding row with number {number}: {e}")
        return None, None

def get_text_from_next_column(row_element, cell_index):
    try:
        next_column_xpath = f'./td[{cell_index + 2}]'  # Increment by 2 to get the next cell
        return row_element.find_element(By.XPATH, next_column_xpath).text
    except Exception as e:
        print(f"Error getting text from next column: {e}")
        return None

def wait_for_element_to_be_clickable(driver, xpath):
    try:
        button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        return button
    except Exception as e:
        print(xpath, e)
 
def click_button(driver, xpath):
    try:
        button = wait_for_element_to_be_clickable(driver, xpath)
        if button:
            button.click()
    except Exception as e:
        print(xpath, e)
 
def input_text(driver, element_id, text):
    try:
        # Find the email input field by its ID
        email_input = driver.find_element(By.ID, element_id)
        # Clear the input field if needed
        email_input.clear()
        # Input the email address
        email_input.send_keys(text)
        email_input.send_keys(Keys.ENTER)
    except Exception as e:
        print(f"An error occurred: {e}")

def get_text_from_xpath(driver, xpath):
    try:
        element = driver.find_element(By.XPATH, xpath)
        return element.text
    except Exception as e:
        print(f"Error finding element with xpath {xpath}: {e}")
        return None 
    
def find_all_matching_rows_and_get_data(driver, number, table_xpath, specific_text):
    try:
        number_str = str(number).strip()  # Ensure the number is treated as a string
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        matched_data = []
        for row_index, row in enumerate(rows, start=1):
            site_number_text = row.find_element(By.XPATH, f"{table_xpath}/tbody/tr[{row_index}]/td[3]").text.strip()
            specific_column_text = row.find_element(By.XPATH, f"{table_xpath}/tbody/tr[{row_index}]/td[5]").text.strip()
            if site_number_text == number_str and specific_column_text == specific_text:  # Compare as strings
                cell_text_1 = site_number_text
                cell_text_2 = row.find_element(By.XPATH, f"{table_xpath}/tbody/tr[{row_index}]/td[4]/a").text.strip()
                matched_data.append((cell_text_1, cell_text_2))
        return matched_data
    except Exception as e:
        print(f"Error finding matching rows: {e}")
        return []

table_xpath = "//table[@id='ctl00_cphMain_rcUserAdmin_UserAdminControl1_rgrdUsers_ctl00']"
    
def input_number_and_search(driver, search_box_xpath, number):
    try:
        search_box = driver.find_element(By.XPATH, search_box_xpath)
        search_box.clear()
        search_box.send_keys(number)
        search_box.send_keys(Keys.ENTER)
        
        # More robust wait: wait until the table updates with the search results
        WebDriverWait(driver, 10).until(
            EC.text_to_be_present_in_element((By.XPATH, table_xpath + "/tbody/tr[1]/td[3]"), number.strip())
        )
    except Exception as e:
        print(f"Error inputting number in search box: {e}")

def find_matching_row_and_get_data(driver, number, table_xpath):
    try:
        number_str = str(number).strip()  # Ensure the number is treated as a string
        rows = driver.find_elements(By.XPATH, f"{table_xpath}/tbody/tr")
        for row in rows:
            cell_text = row.find_element(By.XPATH, "./td[3]").text.strip()
            if cell_text == number_str:  # Compare as strings
                cell_text_1 = cell_text
                cell_text_2 = row.find_element(By.XPATH, "./td[4]/a").text.strip()
                return cell_text_1, cell_text_2
        return None, None
    except Exception as e:
        print(f"Error finding matching row: {e}")
        return None, None

# Initialize the Chrome driver
# chrome_options=Options()
# chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))#,options=chrome_options
url = 'https://trust.wachter.com/Admin/UserAdmin.aspx'
employee_dropdown_xpath = '//*[@id="ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input"]'
rcb_list_xpath = '/html/body/form/div[1]/div/div/ul'
email_element_id = 'i0116'
dropdown_input_id = 'ctl00_cphMain_RoundContainer1_MassEditControl1_cboEmployees_Input'
driver.get(url)
login_button_xpath = '//*[@id="slide-5-layer-8"]'
click_button(driver, login_button_xpath)
wachter_login_button_xpath = '//*[@id="ctl00_cphMain_ssoLoginControl_btnSingleSignOnLogin"]'
click_button(driver, wachter_login_button_xpath)
next_button_path = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, next_button_path)
email_address = 
input_text(driver, email_element_id, email_address)
sign_in_button_xpath = '//*[@id="idSIButton9"]'
wait_for_element_to_be_clickable(driver, sign_in_button_xpath)
password_element_id = 'i0118'
password = 
input_text(driver, password_element_id, password)
stay_signed_in_button_xpath = '//*[@id="idSIButton9"]'
click_button(driver, stay_signed_in_button_xpath)
time.sleep(7)
def read_excel_sheet(excel_path):
    print(f"Reading Excel file: {excel_path}")  # Diagnostic print
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        names = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            names.append((row[0], row[1]))
        print("Names read from Excel:", names)
        return names
    except Exception as e:
        print(f"Error reading Excel file: {e}")

# Function to save the data to an Excel file
def save_data_to_excel(data, filename):
    workbook = Workbook()
    sheet = workbook.active
    headers = ['Username', 'Full Name', 'Email', 'User Status', 'User Type', 'Customer', 'City', 'State']
    sheet.append(headers)

    for row_data in data:
        sheet.append(row_data)

    workbook.save(filename)

    
def enter_and_process_names(driver, names):
    for first_name, last_name in names:
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_txtFirstName")))
            first_name_field = driver.find_element(By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_txtFirstName")
            first_name_field.clear()
            first_name_field.send_keys(first_name)
            print(f"Entering first name: {first_name}")  # Debug print
            
            last_name_field = driver.find_element(By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_txtLastName")
            last_name_field.clear()
            last_name_field.send_keys(last_name)
            print(f"Entering last name: {last_name}")  # Debug print
            last_name_field.send_keys(Keys.ENTER)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_grdUsers")))
            print(f"Entered and submitted: {first_name} {last_name}")  # Debug print
        except Exception as e:
            print(f"Failed to enter name {first_name} {last_name} due to: {e}")
        time.sleep(3)
        process_active_users(driver)
            
def process_active_users(driver):
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "ctl00_cphMain_rcUserAdmin_UserAdminControl1_rgrdUsers_ctl00"))
        )
        print("Table is present.")

        rows = driver.find_elements(By.CSS_SELECTOR, "#ctl00_cphMain_rcUserAdmin_UserAdminControl1_rgrdUsers_ctl00 tbody tr")
        print(f"Found {len(rows)} rows in the table.")
        active_user_data = []
        for row in rows:
            columns = row.find_elements(By.TAG_NAME, "td")
            status = columns[4].text.strip()  # Assuming the status is in the 5th column
            if status == "Active":
                # Capture the relevant data from each column
                user_data = [col.text.strip() for col in columns[1:9]]  # Adjust indices if needed
                print(f"Active user found: {user_data}")
                active_user_data.append(user_data)

                # Find the pencil icon in the first cell and click it
                edit_icon = columns[0].find_element(By.CSS_SELECTOR, "a[id*='hypEdit']")
                edit_icon.click()
                print(f"Clicked the pencil icon for: {user_data[0]}")  # Assuming first item is the username
                hr_details_tab = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='HR Details']"))
                )
                hr_details_tab.click()
                
                # Wait for the "HR Details" section to load
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.ID, "ctl00_ctl00_cphMain_cphMain_rcProfileContent_ProfileMgmtIndividualInfo1_dlSupervisorUser"))
                )

                # Capture the 'Supervisor' value
                supervisor_value = driver.find_element(By.ID, "ctl00_ctl00_cphMain_cphMain_rcProfileContent_ProfileMgmtIndividualInfo1_dlSupervisorUser").text.strip()

                print(f"Supervisor: {supervisor_value}")
                return supervisor_value
                # Add a wait condition here if needed, e.g., wait for a modal or new page to load after clicking the pencil icon
                
        return active_user_data
    except TimeoutException:
        print("Timed out waiting for the user table to load.")
    except NoSuchElementException:
        print("Could not find the user table element on the page.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Initialize WebDriver and load the webpage
extracted_data = [
    # ... your tuples of extracted data
]

# Now, save the extracted data to a new Excel file

# Read names from an Excel sheet
excel_path = r"C:\Users\beckett.mcfarland\Documents\Names.xlsx"
names = read_excel_sheet(excel_path)

# Enter each name into the webpage and process active users
enter_and_process_names(driver, names)

save_data_to_excel(extracted_data, r"C:\Users\beckett.mcfarland\Documents\karen_output.xlsx")
# Close the driver
driver.quit()
