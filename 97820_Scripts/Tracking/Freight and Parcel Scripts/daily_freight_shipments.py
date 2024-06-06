import openpyxl

# Paths of the workbooks
source_workbook_path = r"l:\Warehouse\Freight shipping.xlsx"
destination_workbook_path = r"C:\Users\beckett.mcfarland\Desktop\Beckett-McFarland\97820_Scripts\Tracking\Outputs\Freight Shipments.xlsx"

# Load the workbooks
source_workbook = openpyxl.load_workbook(source_workbook_path, data_only=True, read_only=True)
destination_workbook = openpyxl.load_workbook(destination_workbook_path)

# Specify the sheets to work with
source_sheet = source_workbook['2024']
destination_sheet = destination_workbook['Sheet1']

# Clear existing data in the destination sheet, assuming row 1 has headers
for row in destination_sheet.iter_rows(min_row=2, max_row=destination_sheet.max_row):
    for cell in row:
        cell.value = None

# Filter value and column (adjust these as necessary)
filter_value = 97820  # Assuming the value is numeric
filter_column_index = 4  # Adjust as necessary, 9 for the 9th column

print(f"Filtering for rows with value '{filter_value}' in column {filter_column_index}")

# Count of matched rows for debugging
matched_rows = 0

# Row index to start writing data in the destination sheet
row_index = 2  # Assuming row 1 has headers

# Iterate over the rows in the source sheet
for row in source_sheet.iter_rows(min_row=2, max_col=source_sheet.max_column, values_only=True):  # Assuming row 1 is headers
    # Check if the row meets the filter criterion
    cell_value = row[filter_column_index - 1]
    # Convert to string and strip whitespaces for a fair comparison
    if isinstance(cell_value, str):
        cell_value = cell_value.strip()
    if cell_value == filter_value:
        matched_rows += 1
        # Write directly to the destination sheet
        for col_index, value in enumerate(row, start=1):
            destination_sheet.cell(row=row_index, column=col_index, value=value)
        row_index += 1

print(f"Total matched rows: {matched_rows}")

# Save the workbook
if matched_rows > 0:
    destination_workbook.save(destination_workbook_path)
    print(f"Filtered data has been successfully written and saved at {destination_workbook_path}.")
else:
    print("No rows matched the filter criteria. No data written.")
